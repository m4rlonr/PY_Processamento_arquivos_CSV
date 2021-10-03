from openpyxl import load_workbook

print("Carregando dados")
pagina = load_workbook('sabado.xlsx')['7_sabado']
print("Dados carregados")

contador = 0
pivo = 0
quantidades = []
print("váriaveis criadas")

print("inicio processamento")
for item in pagina.iter_rows(min_row=2):
    print(item[0].value)
    data = item[1].value
    if(type(data) == str):
        hora = int(data.split(" ")[1].split(":")[0])
    else:
        quantidades.append({"Horas": pivo, "Quantidade": contador})
    if(pivo == hora):
        contador += 1
    else:
        quantidades.append({"Horas": pivo, "Quantidade": contador})
        pivo = hora
        contador = 0

print(quantidades)
